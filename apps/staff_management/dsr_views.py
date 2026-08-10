"""DSR seller-code allocation API (Administration module).

Endpoints:
* ``GET  dsr-sales-codes/``            — list/search existing allocations.
* ``GET  dsr-sales-codes/lookup/?pf_number=`` — is this PF already allocated? If
  not, autofilled name/branch/department + the next code that would be issued.
* ``POST dsr-sales-codes/allocate/``   — allocate a code to a PF (idempotent per PF).
* ``POST dsr-sales-codes/upload-csv/`` — bulk import the listing (upsert on PF).

The code-generation and PF rules live in :mod:`apps.staff_management.dsr`.
"""

import csv
import io
from datetime import datetime

from django.db import IntegrityError, transaction
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters, generics, serializers, status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from core.pagination import StandardPagination

from .dsr import autofill_from_roster, next_sales_code
from .models import DSRSalesCode


class DSRSalesCodeSerializer(serializers.ModelSerializer):
    class Meta:
        model = DSRSalesCode
        fields = "__all__"
        # sales_code is system-generated on allocate; never accept it from the body.
        read_only_fields = ["sales_code", "created_at", "updated_at"]


def _is_admin(user) -> bool:
    """Only administrators may allocate codes or bulk-import the listing."""
    return bool(
        user
        and user.is_authenticated
        and (user.is_superuser or user.is_staff or user.groups.filter(name="staff_mgt").exists())
    )


def _parse_date(value):
    if not value:
        return None
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S", "%d-%b-%Y"):
        try:
            return datetime.strptime(text[:19], fmt).date()
        except ValueError:
            continue
    return None


class DSRSalesCodeListView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = DSRSalesCodeSerializer
    pagination_class = StandardPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["branch", "role", "department"]
    search_fields = ["pf_number", "sales_code", "salesperson", "branch", "role", "team_leader"]
    ordering_fields = ["sales_code", "salesperson", "branch", "created_at"]
    queryset = DSRSalesCode.objects.all()


class DSRSalesCodeLookupView(APIView):
    """Check a PF before allocating: existing code, or an autofilled preview."""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        pf = (request.query_params.get("pf_number") or "").strip()
        if not pf:
            return Response({"detail": "pf_number is required."}, status=status.HTTP_400_BAD_REQUEST)

        existing = DSRSalesCode.objects.filter(pf_number=pf).first()
        if existing:
            return Response({
                "allocated": True,
                "record": DSRSalesCodeSerializer(existing).data,
            })

        return Response({
            "allocated": False,
            "next_sales_code": next_sales_code(),
            "suggested": {"pf_number": pf, **autofill_from_roster(pf)},
        })


class DSRSalesCodeAllocateView(APIView):
    """Allocate the next DSR code to a PF. Returns the existing one if already set."""

    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not _is_admin(request.user):
            return Response({"detail": "Administrators only."}, status=status.HTTP_403_FORBIDDEN)

        pf = str(request.data.get("pf_number") or "").strip()
        if not pf:
            return Response({"pf_number": "This field is required."}, status=status.HTTP_400_BAD_REQUEST)

        existing = DSRSalesCode.objects.filter(pf_number=pf).first()
        if existing:
            return Response({
                "already_allocated": True,
                "detail": f"PF {pf} already has sales code {existing.sales_code}.",
                "record": DSRSalesCodeSerializer(existing).data,
            }, status=status.HTTP_200_OK)

        payload = {
            "pf_number": pf,
            "salesperson": str(request.data.get("salesperson") or "").strip(),
            "branch": str(request.data.get("branch") or "").strip(),
            "department": str(request.data.get("department") or "").strip(),
            "role": str(request.data.get("role") or "").strip(),
            "team_leader": str(request.data.get("team_leader") or "").strip(),
            "date_of_employment": _parse_date(request.data.get("date_of_employment")),
            "allocation_date": _parse_date(request.data.get("allocation_date")) or datetime.now().date(),
        }

        # Generate + save atomically; retry once if another allocation grabbed the
        # same number between read and write (unique constraint is the backstop).
        for _ in range(5):
            code = next_sales_code()
            try:
                with transaction.atomic():
                    record = DSRSalesCode.objects.create(sales_code=code, **payload)
                return Response({
                    "already_allocated": False,
                    "record": DSRSalesCodeSerializer(record).data,
                }, status=status.HTTP_201_CREATED)
            except IntegrityError:
                # PF race → someone allocated this PF first; return theirs.
                dup = DSRSalesCode.objects.filter(pf_number=pf).first()
                if dup:
                    return Response({
                        "already_allocated": True,
                        "detail": f"PF {pf} already has sales code {dup.sales_code}.",
                        "record": DSRSalesCodeSerializer(dup).data,
                    }, status=status.HTTP_200_OK)
                # else sales_code race → loop and try the next number.
                continue

        return Response(
            {"detail": "Could not allocate a unique code, please retry."},
            status=status.HTTP_409_CONFLICT,
        )


def _clean_cell(value) -> str:
    """Stringify a cell, trimming Excel's numeric ``.0`` tails (e.g. PF 3804.0)."""
    if value is None:
        return ""
    s = str(value).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def _read_upload_rows(upload):
    """Return a list of rows (each a list of string cells) from a CSV or XLSX upload.

    Handles both formats so the listing can be uploaded exactly as exported — no
    need to strip the Excel's blank header row or leading empty columns first.
    """
    name = (getattr(upload, "name", "") or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        import openpyxl

        wb = openpyxl.load_workbook(upload, read_only=True, data_only=True)
        ws = wb["List"] if "List" in wb.sheetnames else wb[wb.sheetnames[0]]
        return [[_clean_cell(c) for c in row] for row in ws.iter_rows(values_only=True)]

    raw = upload.read()
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin-1")
    return [[_clean_cell(c) for c in row] for row in csv.reader(io.StringIO(text))]


class DSRSalesCodeCSVUploadView(APIView):
    """Bulk import the DSR listing (CSV **or** XLSX). Upserts on PF number.

    The header row is located by name (case-insensitive), so a blank leading row,
    empty leading columns, and extra columns (CH, unnamed) in the export are all
    tolerated. Recognised headers: PF_NO, SALESCODE, SALESPERSON, Branch, Role,
    Team Leader, DATE OF EMPLOYMENT, ALLOCATION DATE.
    """

    permission_classes = [IsAuthenticated]

    HEADER_MAP = {
        "pf_no": "pf_number", "pf no": "pf_number", "pf_number": "pf_number", "pf": "pf_number",
        "salescode": "sales_code", "sales code": "sales_code", "sales_code": "sales_code",
        "salesperson": "salesperson", "sales person": "salesperson", "name": "salesperson",
        "branch": "branch",
        "role": "role",
        "department": "department",
        "team leader": "team_leader", "team_leader": "team_leader",
        "date of employment": "date_of_employment", "date_of_employment": "date_of_employment",
        "allocation date": "allocation_date", "allocation_date": "allocation_date",
    }

    def post(self, request):
        if not _is_admin(request.user):
            return Response({"detail": "Administrators only."}, status=status.HTTP_403_FORBIDDEN)

        upload = request.FILES.get("file")
        if not upload:
            return Response({"detail": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            rows = _read_upload_rows(upload)
        except Exception as exc:
            return Response({"detail": f"Could not read the file: {exc}"}, status=status.HTTP_400_BAD_REQUEST)

        # Locate the header row: the first row that maps to both PF and sales code.
        col_map = {}
        header_idx = None
        for i, row in enumerate(rows):
            mapping = {j: self.HEADER_MAP[c.strip().lower()]
                       for j, c in enumerate(row) if c.strip().lower() in self.HEADER_MAP}
            if "pf_number" in mapping.values() and "sales_code" in mapping.values():
                col_map, header_idx = mapping, i
                break

        if header_idx is None:
            return Response(
                {"detail": "Could not find the header row — need PF_NO and SALESCODE columns."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created = updated = skipped = 0
        errors = []
        with transaction.atomic():
            for n, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
                data = {field: (row[j] if j < len(row) else "") for j, field in col_map.items()}
                pf = data.get("pf_number", "").strip()
                code = data.get("sales_code", "").strip()
                if not pf or not code:
                    skipped += 1
                    continue

                fields = {
                    "sales_code": code,
                    "salesperson": data.get("salesperson", ""),
                    "branch": data.get("branch", ""),
                    "department": data.get("department", ""),
                    "role": data.get("role", ""),
                    "team_leader": data.get("team_leader", ""),
                    "date_of_employment": _parse_date(data.get("date_of_employment")),
                    "allocation_date": _parse_date(data.get("allocation_date")),
                }
                try:
                    with transaction.atomic():
                        _, was_created = DSRSalesCode.objects.update_or_create(
                            pf_number=pf, defaults=fields,
                        )
                    created += int(was_created)
                    updated += int(not was_created)
                except IntegrityError:
                    errors.append(f"Row {n}: sales code {code} already used by another PF.")
                    skipped += 1

        return Response({
            "created": created, "updated": updated, "skipped": skipped,
            "errors": errors[:50],
        }, status=status.HTTP_200_OK)
